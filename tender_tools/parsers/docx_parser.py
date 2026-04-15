"""DOCX-парсер: извлекает текст, таблицы и изображения в ParsedDocument.

На базе экспериментального parse_docx.py, адаптирован под унифицированную
JSON-схему с поддержкой многоуровневых заголовков таблиц, row_type,
colspan/rowspan и извлечения изображений.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

from docx import Document
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph

from .schema import (
    CellObject,
    ImageBlock,
    ParsedDocument,
    TableBlock,
    TableRow,
    TextBlock,
)

logger = logging.getLogger(__name__)

# Паттерн для определения номера раздела в тексте (1., 1.1, 2.3.1 и т.д.)
_SECTION_RE = re.compile(r"^(\d+(?:\.\d+)*)[\.\s\)]")


# ---------------------------------------------------------------------------
# Утилиты
# ---------------------------------------------------------------------------

def _clean_cell_text(text: str) -> str:
    """Очищает текст ячейки от лишних переносов и пробелов."""
    text = text.replace("\n", " ").replace("\r", " ")
    while "  " in text:
        text = text.replace("  ", " ")
    return text.strip()


def _extract_heading_level(para: Paragraph) -> int | None:
    """Возвращает уровень заголовка (1-6) или None для обычного текста."""
    style_name = (para.style.name or "").lower()
    if style_name.startswith("heading"):
        try:
            level = int(style_name.replace("heading", "").strip())
            return min(max(level, 1), 6)
        except ValueError:
            pass
    return None


def _extract_section_path(text: str) -> str | None:
    """Извлекает номер раздела из начала текста: '1.1 Общие сведения' → '1.1'."""
    m = _SECTION_RE.match(text.strip())
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# Обход элементов документа в порядке следования
# ---------------------------------------------------------------------------

def _iter_block_elements(doc: Document):
    """Итерирует по элементам тела документа, возвращая Paragraph или Table."""
    for element in doc.element.body:
        tag = element.tag
        if tag == qn("w:p"):
            yield Paragraph(element, doc)
        elif tag == qn("w:tbl"):
            yield Table(element, doc)


# ---------------------------------------------------------------------------
# Извлечение таблицы с поддержкой colspan и row_type
# ---------------------------------------------------------------------------

def _get_grid_span(tc_element) -> int:
    """Получает colspan ячейки из XML (gridSpan)."""
    grid_span = tc_element.find(qn("w:tcPr"))
    if grid_span is not None:
        gs = grid_span.find(qn("w:gridSpan"))
        if gs is not None:
            val = gs.get(qn("w:val"))
            if val and val.isdigit():
                return int(val)
    return 1


def _is_vertically_merged(tc_element) -> str | None:
    """Проверяет вертикальное объединение ячейки.

    Возвращает:
        'restart' — начало объединения
        'continue' — продолжение
        None — не объединена
    """
    tc_pr = tc_element.find(qn("w:tcPr"))
    if tc_pr is not None:
        v_merge = tc_pr.find(qn("w:vMerge"))
        if v_merge is not None:
            val = v_merge.get(qn("w:val"))
            return val if val else "continue"
    return None


def _classify_row(cells_text: list[str], col_count: int, row_idx: int) -> str:
    """Эвристика для определения типа строки.

    - header: первая строка или строка, где все ячейки — короткие заголовки
    - section: строка-разделитель (одна непустая ячейка на всю ширину)
    - data: обычная строка данных
    """
    non_empty = [c for c in cells_text if c.strip()]

    if row_idx == 0:
        return "header"

    if len(non_empty) <= 2 and col_count >= 4:
        total_len = sum(len(c) for c in non_empty)
        if total_len > 0 and all(len(c) < 200 for c in non_empty):
            return "section"

    return "data"


def _find_nested_tables(tc_element) -> list:
    """Находит вложенные <w:tbl> внутри ячейки."""
    return tc_element.findall(qn("w:tbl"))


def _extract_cell_text_deep(tc_element) -> str:
    """Извлекает текст из ячейки, включая вложенные таблицы.

    python-docx cell.text видит только прямые <w:p>, но не вложенные
    <w:tbl>. Эта функция рекурсивно обходит всё содержимое.
    """
    texts: list[str] = []
    for elem in tc_element.iter():
        if elem.tag == qn("w:t") and elem.text:
            texts.append(elem.text)
    return _clean_cell_text(" ".join(texts))


def _is_layout_wrapper(table: Table) -> bool:
    """Определяет, является ли таблица layout-обёрткой (невидимой оболочкой).

    Признаки layout-таблицы:
    - Все границы ячеек = nil (невидимы)
    - Содержит вложенные таблицы <w:tbl> внутри ячеек
    - Мало уникальных ячеек с текстом в прямых <w:p>
    """
    if len(table.rows) == 0:
        return False

    # Проверяем первую строку
    first_row = table.rows[0]
    tr = first_row._tr
    tc_elements = tr.findall(qn("w:tc"))

    if not tc_elements:
        return False

    # Если первая ячейка содержит вложенную таблицу — это обёртка
    nested = tc_elements[0].findall(qn("w:tbl"))
    if nested:
        return True

    return False


def _extract_nested_tables(
    table: Table, doc_id: str, table_idx_start: int
) -> list[TableBlock]:
    """Извлекает все вложенные таблицы из layout-обёртки.

    Обходит каждую ячейку внешней таблицы, находит вложенные <w:tbl>,
    и извлекает из них данные как из обычных таблиц.
    """
    from docx.table import Table as DocxTable

    nested_blocks: list[TableBlock] = []
    table_counter = table_idx_start

    for row in table.rows:
        tr = row._tr
        for tc in tr.findall(qn("w:tc")):
            for nested_tbl in tc.findall(qn("w:tbl")):
                nested_table = DocxTable(nested_tbl, table._parent)
                table_counter += 1
                block = _extract_table_direct(
                    nested_table, doc_id, table_counter
                )
                if block.rows:
                    nested_blocks.append(block)

    return nested_blocks


def _extract_table(table: Table, doc_id: str, table_idx: int) -> TableBlock | list[TableBlock]:
    """Извлекает таблицу в TableBlock.

    Если таблица — layout-обёртка с вложенными таблицами, возвращает
    список TableBlock из вложенных. Иначе — один TableBlock.
    """
    if _is_layout_wrapper(table):
        logger.debug("Обнаружена layout-обёртка, извлекаю вложенные таблицы")
        nested = _extract_nested_tables(table, doc_id, table_idx)
        if nested:
            return nested
        # Если вложенных таблиц нет — пробуем как обычную
        logger.debug("Вложенных таблиц не нашлось, парсим как обычную")

    return _extract_table_direct(table, doc_id, table_idx)


def _extract_table_direct(table: Table, doc_id: str, table_idx: int) -> TableBlock:
    """Извлекает таблицу в TableBlock с поддержкой colspan и row_type."""
    rows_data: list[TableRow] = []
    col_count = 0

    for row_idx, row in enumerate(table.rows):
        cells = []
        cells_text: list[str] = []

        prev_tc = None
        for cell in row.cells:
            tc = cell._tc
            if tc is prev_tc:
                continue
            prev_tc = tc

            # Используем глубокое извлечение текста (для вложенных структур)
            text = _extract_cell_text_deep(tc)
            cells_text.append(text)
            colspan = _get_grid_span(tc)

            if colspan > 1:
                cells.append(CellObject(value=text, colspan=colspan))
            else:
                cells.append(text)

        logical_width = sum(
            c.colspan if isinstance(c, CellObject) else 1 for c in cells
        )
        col_count = max(col_count, logical_width)

        row_type = _classify_row(cells_text, logical_width, row_idx)
        rows_data.append(TableRow(row_type=row_type, cells=cells))

    # Фильтруем полностью пустые строки
    rows_data = [
        r for r in rows_data
        if any(
            (c.value if isinstance(c, CellObject) else c).strip()
            for c in r.cells
        )
    ]

    block_id = f"{doc_id}_tbl_{table_idx:03d}"
    return TableBlock(
        block_id=block_id,
        col_count=col_count,
        rows=rows_data,
    )


# ---------------------------------------------------------------------------
# Извлечение изображений
# ---------------------------------------------------------------------------

def _extract_images(doc: Document, doc_id: str, images_dir: Path) -> dict[str, ImageBlock]:
    """Извлекает все встроенные изображения из DOCX.

    Возвращает словарь {relationship_id: ImageBlock} для привязки
    к позиции в документе.
    """
    images: dict[str, ImageBlock] = {}
    img_counter = 0

    for rel_id, rel in doc.part.rels.items():
        if "image" in rel.reltype:
            try:
                image_part = rel.target_part
                content_type = image_part.content_type
                ext = _content_type_to_ext(content_type)

                img_counter += 1
                img_filename = f"{doc_id}_img_{img_counter:03d}{ext}"
                img_path = images_dir / img_filename

                images_dir.mkdir(parents=True, exist_ok=True)
                img_path.write_bytes(image_part.blob)

                images[rel_id] = ImageBlock(
                    block_id=f"{doc_id}_img_{img_counter:03d}",
                    image_ref=f"images/{img_filename}",
                    ocr_status="pending",
                )
                logger.debug("Извлечено изображение: %s", img_filename)
            except Exception as exc:
                logger.warning("Не удалось извлечь изображение %s: %s", rel_id, exc)

    return images


def _content_type_to_ext(content_type: str) -> str:
    """Маппинг MIME-типа в расширение файла."""
    mapping = {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/gif": ".gif",
        "image/bmp": ".bmp",
        "image/tiff": ".tiff",
        "image/x-emf": ".emf",
        "image/x-wmf": ".wmf",
    }
    return mapping.get(content_type, ".png")


# ---------------------------------------------------------------------------
# Основной парсер
# ---------------------------------------------------------------------------

def parse_docx(file_path: Path, doc_id: str, output_dir: Path | None = None) -> ParsedDocument:
    """Парсит DOCX-файл в ParsedDocument.

    Args:
        file_path: путь к .docx файлу.
        doc_id: идентификатор документа (doc_001, doc_002, ...).
        output_dir: директория для сохранения артефактов (images/, tables/).
                    Если None — изображения не извлекаются на диск.

    Returns:
        ParsedDocument с блоками текста, таблиц и изображений.
    """
    logger.info("Парсинг DOCX: %s", file_path.name)
    doc = Document(str(file_path))

    blocks = []
    block_counter = 0
    table_counter = 0
    current_section: str | None = None

    # Извлекаем изображения (если есть output_dir)
    image_blocks: dict[str, ImageBlock] = {}
    if output_dir:
        images_dir = output_dir / doc_id / "images"
        image_blocks = _extract_images(doc, doc_id, images_dir)

    for element in _iter_block_elements(doc):

        if isinstance(element, Paragraph):
            text = element.text.strip()
            if not text:
                continue

            block_counter += 1
            heading_level = _extract_heading_level(element)
            section_path = _extract_section_path(text)

            if section_path:
                current_section = section_path
            elif heading_level:
                current_section = None

            blocks.append(TextBlock(
                block_id=f"{doc_id}_b{block_counter:03d}",
                content=text,
                heading_level=heading_level,
                section_path=section_path or current_section,
            ))

            # Проверяем наличие inline-изображений в параграфе
            for run in element.runs:
                drawing_elements = run._element.findall(f".//{qn('a:blip')}")
                for blip in drawing_elements:
                    embed_id = blip.get(qn("r:embed"))
                    if embed_id and embed_id in image_blocks:
                        img_block = image_blocks.pop(embed_id)
                        img_block.section_path = current_section
                        blocks.append(img_block)

        elif isinstance(element, Table):
            table_counter += 1
            result = _extract_table(element, doc_id, table_counter)

            # _extract_table может вернуть список (для layout-обёрток)
            table_blocks_to_add = result if isinstance(result, list) else [result]

            for tb in table_blocks_to_add:
                tb.section_path = current_section

                if output_dir:
                    tables_dir = output_dir / doc_id / "tables"
                    tables_dir.mkdir(parents=True, exist_ok=True)
                    xlsx_name = f"{tb.block_id}.xlsx"
                    xlsx_path = tables_dir / xlsx_name
                    _export_table_to_xlsx(tb, xlsx_path)
                    tb.xlsx_ref = f"tables/{xlsx_name}"

                blocks.append(tb)

            if isinstance(result, list):
                table_counter += len(result) - 1

    # Добавляем оставшиеся изображения (не привязанные к параграфам)
    for img_block in image_blocks.values():
        blocks.append(img_block)

    parsed = ParsedDocument(
        doc_id=doc_id,
        source_filename=file_path.name,
        source_format="docx",
        blocks=blocks,
    )

    logger.info(
        "DOCX распарсен: %d текст, %d таблиц, %d изображений",
        len(parsed.text_blocks),
        len(parsed.table_blocks),
        len(parsed.image_blocks),
    )
    return parsed


# ---------------------------------------------------------------------------
# Экспорт таблицы в XLSX
# ---------------------------------------------------------------------------

def _export_table_to_xlsx(table_block: TableBlock, xlsx_path: Path) -> None:
    """Экспортирует TableBlock в отформатированный XLSX-файл."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица"

    for row_idx, table_row in enumerate(table_block.rows):
        col_offset = 1
        for cell_data in table_row.cells:
            if isinstance(cell_data, CellObject):
                value = cell_data.value
                colspan = cell_data.colspan
            else:
                value = cell_data
                colspan = 1

            cell = ws.cell(row=row_idx + 1, column=col_offset, value=value)
            cell.border = thin_border

            if table_row.row_type == "header":
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            elif table_row.row_type == "section":
                cell.font = Font(name="Calibri", bold=True, size=10)
                cell.alignment = data_align
            else:
                cell.font = data_font
                cell.alignment = data_align
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

            if colspan > 1:
                end_col = col_offset + colspan - 1
                ws.merge_cells(
                    start_row=row_idx + 1, start_column=col_offset,
                    end_row=row_idx + 1, end_column=end_col,
                )
                for c in range(col_offset + 1, end_col + 1):
                    ws.cell(row=row_idx + 1, column=c).border = thin_border

            col_offset += colspan

    for col_idx in range(1, table_block.col_count + 1):
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    text = str(cell.value)
                    words = text.split()
                    longest = max((len(w) for w in words), default=0)
                    max_len = max(max_len, min(max(longest, len(text) // 3), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    wb.save(str(xlsx_path))
