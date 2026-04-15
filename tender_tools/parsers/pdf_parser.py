"""PDF-парсер: извлекает текст, таблицы и изображения в ParsedDocument.

На базе экспериментального parse_pdf.py с поддержкой:
- Склейки многостраничных таблиц
- Обнаружения сканов (страница-изображение без текста)
- Извлечения изображений
- Определения section_path
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pdfplumber

from .schema import (
    CellObject,
    ImageBlock,
    ParsedDocument,
    TableBlock,
    TableRow,
    TextBlock,
)

logger = logging.getLogger(__name__)

_SECTION_RE = re.compile(r"^(\d+(?:\.\d+)*)[\.\s\)]")

# Минимальная длина текста страницы, ниже которой считаем её сканом
_SCAN_TEXT_THRESHOLD = 50


def _extract_section_path(text: str) -> str | None:
    """Извлекает номер раздела из начала текста: '1.1 Общие сведения' → '1.1'."""
    m = _SECTION_RE.match(text.strip())
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# Утилиты очистки
# ---------------------------------------------------------------------------

def _clean_cell(value: str | None) -> str:
    if value is None:
        return ""
    text = value.replace("\n", " ").replace("\r", " ")
    while "  " in text:
        text = text.replace("  ", " ")
    return text.strip()


def _clean_table(raw_table: list[list[str | None]]) -> list[list[str]]:
    return [[_clean_cell(cell) for cell in row] for row in raw_table]


# ---------------------------------------------------------------------------
# Извлечение данных со страницы
# ---------------------------------------------------------------------------

def _extract_page_data(page: pdfplumber.page.Page) -> dict:
    """Извлекает таблицы и текст вне таблиц со страницы."""
    result = {"tables": [], "text_blocks": []}

    tables = page.find_tables()
    table_bboxes = [t.bbox for t in tables]

    for table in tables:
        raw_rows = table.extract()
        if raw_rows:
            cleaned = _clean_table(raw_rows)
            result["tables"].append({
                "rows": cleaned,
                "bbox": table.bbox,
            })

    page_bbox = page.bbox
    filtered_page = page
    for bbox in table_bboxes:
        clamped = (
            max(bbox[0], page_bbox[0]),
            max(bbox[1], page_bbox[1]),
            min(bbox[2], page_bbox[2]),
            min(bbox[3], page_bbox[3]),
        )
        if clamped[2] <= clamped[0] or clamped[3] <= clamped[1]:
            continue
        try:
            filtered_page = filtered_page.outside_bbox(clamped)
        except ValueError:
            continue

    text = filtered_page.extract_text()
    if text and text.strip():
        words = filtered_page.extract_words()
        top_pos = min(w["top"] for w in words) if words else 0
        result["text_blocks"].append({
            "content": text.strip(),
            "top": top_pos,
        })

    return result


# ---------------------------------------------------------------------------
# Детекция сканов
# ---------------------------------------------------------------------------

def _is_scan_page(page: pdfplumber.page.Page) -> bool:
    """Определяет, является ли страница сканом (изображением без текста)."""
    text = page.extract_text() or ""
    if len(text.strip()) >= _SCAN_TEXT_THRESHOLD:
        return False
    images = page.images
    if not images:
        return False
    page_area = page.width * page.height
    total_img_area = sum(
        abs(img.get("x1", 0) - img.get("x0", 0)) * abs(img.get("bottom", 0) - img.get("top", 0))
        for img in images
    )
    return total_img_area > page_area * 0.5


# ---------------------------------------------------------------------------
# Склейка многостраничных таблиц
# ---------------------------------------------------------------------------

def _rows_are_similar(row_a: list[str], row_b: list[str]) -> bool:
    if len(row_a) != len(row_b):
        return False
    return all(a.strip() == b.strip() for a, b in zip(row_a, row_b))


def _try_merge_split_row(last_row: list[str], first_row: list[str]) -> list[str] | None:
    if len(last_row) != len(first_row):
        return None
    merged = []
    can_merge = False
    for a, b in zip(last_row, first_row):
        if a and b:
            merged.append(f"{a} {b}".strip())
            can_merge = True
        elif a:
            merged.append(a)
        elif b:
            merged.append(b)
            can_merge = True
        else:
            merged.append("")
    return merged if can_merge else None


def _classify_pdf_row(cells: list[str], col_count: int, is_first: bool) -> str:
    """Эвристика для определения типа строки в PDF-таблице."""
    non_empty = [c for c in cells if c.strip()]
    if is_first:
        return "header"
    if len(non_empty) <= 2 and col_count >= 4:
        return "section"
    return "data"


def _merge_tables_across_pages(
    pages_data: list[dict],
    doc_id: str,
) -> tuple[list[dict], int]:
    """Склеивает многостраничные таблицы и формирует ordered_elements.

    Возвращает (ordered_elements, table_count).
    """
    ordered_elements: list[dict] = []
    table_counter = 0

    pending_rows: list[list[str]] | None = None
    pending_ncols: int = 0
    pending_start_page: int = 0

    def finalize_pending(end_page: int):
        nonlocal pending_rows, pending_ncols, table_counter
        if pending_rows and len(pending_rows) > 0:
            table_counter += 1
            ordered_elements.append({
                "type": "table",
                "raw_rows": pending_rows,
                "ncols": pending_ncols,
                "page": pending_start_page,
                "page_end": end_page,
                "table_idx": table_counter,
            })
        pending_rows = None
        pending_ncols = 0

    for page_idx, page_data in enumerate(pages_data):
        page_num = page_idx + 1
        tables = page_data["tables"]
        text_blocks = page_data["text_blocks"]

        elements = []
        for tb in text_blocks:
            elements.append(("text", tb["content"], tb["top"]))
        for tbl in tables:
            elements.append(("table", tbl["rows"], tbl["bbox"][1]))
        elements.sort(key=lambda e: e[2])

        for elem_type, content, _ in elements:
            if elem_type == "text":
                finalize_pending(page_num)
                ordered_elements.append({
                    "type": "text",
                    "content": content,
                    "page": page_num,
                })
            elif elem_type == "table":
                curr_rows: list[list[str]] = content
                if not curr_rows:
                    continue
                curr_ncols = len(curr_rows[0])

                if pending_rows is not None and curr_ncols == pending_ncols:
                    start_idx = 0
                    if len(curr_rows) > 1 and _rows_are_similar(pending_rows[0], curr_rows[0]):
                        start_idx = 1
                    rows_to_add = curr_rows[start_idx:]
                    if rows_to_add and pending_rows:
                        merged_row = _try_merge_split_row(pending_rows[-1], rows_to_add[0])
                        if merged_row is not None:
                            pending_rows[-1] = merged_row
                            rows_to_add = rows_to_add[1:]
                    pending_rows.extend(rows_to_add)
                else:
                    finalize_pending(page_num)
                    pending_rows = curr_rows
                    pending_ncols = curr_ncols
                    pending_start_page = page_num

    finalize_pending(len(pages_data))
    return ordered_elements, table_counter


# ---------------------------------------------------------------------------
# Основной парсер
# ---------------------------------------------------------------------------

def parse_pdf(file_path: Path, doc_id: str, output_dir: Path | None = None) -> ParsedDocument:
    """Парсит PDF-файл в ParsedDocument.

    Args:
        file_path: путь к .pdf файлу.
        doc_id: идентификатор документа.
        output_dir: директория для артефактов (images/, tables/).

    Returns:
        ParsedDocument.
    """
    logger.info("Парсинг PDF: %s", file_path.name)

    pages_data: list[dict] = []
    scan_pages: list[int] = []
    total_pages = 0

    with pdfplumber.open(str(file_path)) as pdf:
        total_pages = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            if _is_scan_page(page):
                scan_pages.append(i + 1)
                pages_data.append({"tables": [], "text_blocks": []})
                logger.debug("Страница %d — скан", i + 1)
            else:
                pages_data.append(_extract_page_data(page))

    ordered_elements, table_count = _merge_tables_across_pages(pages_data, doc_id)

    blocks = []
    block_counter = 0
    current_section: str | None = None

    for elem in ordered_elements:
        block_counter += 1

        if elem["type"] == "text":
            content = elem["content"]
            section_path = _extract_section_path(content)
            if section_path:
                current_section = section_path

            blocks.append(TextBlock(
                block_id=f"{doc_id}_b{block_counter:03d}",
                page=elem.get("page"),
                content=content,
                section_path=section_path or current_section,
            ))

        elif elem["type"] == "table":
            raw_rows = elem["raw_rows"]
            ncols = elem["ncols"]
            table_idx = elem["table_idx"]

            table_rows = []
            for row_idx, raw_row in enumerate(raw_rows):
                row_type = _classify_pdf_row(raw_row, ncols, is_first=(row_idx == 0))
                cells = [_clean_cell(c) for c in raw_row]
                table_rows.append(TableRow(row_type=row_type, cells=cells))

            block_id = f"{doc_id}_tbl_{table_idx:03d}"
            table_block = TableBlock(
                block_id=block_id,
                page=elem.get("page"),
                page_end=elem.get("page_end"),
                col_count=ncols,
                rows=table_rows,
                section_path=current_section,
            )

            if output_dir:
                tables_dir = output_dir / doc_id / "tables"
                tables_dir.mkdir(parents=True, exist_ok=True)
                xlsx_name = f"{doc_id}_tbl_{table_idx:03d}.xlsx"
                xlsx_path = tables_dir / xlsx_name
                _export_table_to_xlsx(table_block, xlsx_path)
                table_block.xlsx_ref = f"tables/{xlsx_name}"

            blocks.append(table_block)

    # Добавляем страницы-сканы как ImageBlock
    for scan_page in scan_pages:
        block_counter += 1
        blocks.append(ImageBlock(
            block_id=f"{doc_id}_scan_{scan_page:03d}",
            page=scan_page,
            image_ref=None,
            ocr_status="pending",
        ))

    parsed = ParsedDocument(
        doc_id=doc_id,
        source_filename=file_path.name,
        source_format="pdf",
        total_pages=total_pages,
        blocks=blocks,
    )

    logger.info(
        "PDF распарсен: %d страниц, %d текст, %d таблиц, %d сканов",
        total_pages,
        len(parsed.text_blocks),
        len(parsed.table_blocks),
        len(scan_pages),
    )
    return parsed


# ---------------------------------------------------------------------------
# Экспорт таблицы в XLSX (переиспользуем из docx_parser)
# ---------------------------------------------------------------------------

def _export_table_to_xlsx(table_block: TableBlock, xlsx_path: Path) -> None:
    """Экспортирует TableBlock в XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица"

    for row_idx, table_row in enumerate(table_block.rows):
        for col_idx, cell_data in enumerate(table_row.cells, start=1):
            value = cell_data.value if isinstance(cell_data, CellObject) else cell_data
            cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)
            cell.border = thin_border

            if table_row.row_type == "header":
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            else:
                cell.font = data_font
                cell.alignment = data_align
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    for col_idx in range(1, table_block.col_count + 1):
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    text = str(cell.value)
                    words = text.split()
                    longest = max((len(w) for w in words), default=0)
                    max_len = max(max_len, min(max(longest, len(text) // 3), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    wb.save(str(xlsx_path))
