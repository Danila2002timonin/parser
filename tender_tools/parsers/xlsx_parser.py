"""XLSX-парсер: извлекает таблицы из Excel-файлов в ParsedDocument.

Каждый лист Excel становится отдельным TableBlock.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from .schema import (
    ParsedDocument,
    TableBlock,
    TableRow,
)

logger = logging.getLogger(__name__)


def _clean_cell_value(value) -> str:
    """Приводит значение ячейки Excel к строке."""
    if value is None:
        return ""
    return str(value).strip()


def parse_xlsx(file_path: Path, doc_id: str, output_dir: Path | None = None) -> ParsedDocument:
    """Парсит XLSX-файл в ParsedDocument.

    Каждый лист (sheet) Excel становится отдельным TableBlock.

    Args:
        file_path: путь к .xlsx файлу.
        doc_id: идентификатор документа.
        output_dir: не используется (xlsx — это уже таблица).

    Returns:
        ParsedDocument.
    """
    logger.info("Парсинг XLSX: %s", file_path.name)

    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    blocks = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows_data: list[TableRow] = []
        col_count = 0

        for row_idx, row in enumerate(ws.iter_rows()):
            cells = [_clean_cell_value(cell.value) for cell in row]
            col_count = max(col_count, len(cells))

            if all(c == "" for c in cells):
                continue

            row_type = "header" if row_idx == 0 else "data"
            rows_data.append(TableRow(row_type=row_type, cells=cells))

        if not rows_data:
            continue

        block_id = f"{doc_id}_sheet_{sheet_idx:02d}"
        blocks.append(TableBlock(
            block_id=block_id,
            col_count=col_count,
            rows=rows_data,
        ))

        logger.debug(
            "Лист '%s': %d строк x %d колонок", sheet_name, len(rows_data), col_count
        )

    wb.close()

    parsed = ParsedDocument(
        doc_id=doc_id,
        source_filename=file_path.name,
        source_format="xlsx",
        blocks=blocks,
    )

    logger.info("XLSX распарсен: %d листов → %d таблиц", len(wb.sheetnames), len(blocks))
    return parsed
